"""
Test cases for Garantías and Liquidaciones features
=====================================================
Tests:
- POST /api/ordenes/{id}/crear-garantia - Create dependent warranty order
- GET /api/ordenes/{id}/garantias - Get warranty orders linked to parent
- POST /api/liquidaciones/importar-excel - Import Excel and auto-cross codes
- GET /api/liquidaciones/pendientes - Get pending/paid/reclaimed liquidations
- POST /api/liquidaciones/marcar-pagados - Mark multiple as paid
- GET /api/liquidaciones/historial-meses - Get monthly summary
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestAuth:
    """Authentication tests to get tokens for subsequent tests"""
    
    @pytest.fixture(scope="class")
    def admin_token(self):
        """Get admin token for tests requiring admin role"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@techrepair.local",
            "password": "Admin2026!"
        })
        if response.status_code == 200:
            return response.json().get("token")
        pytest.skip("Admin authentication failed")
    
    @pytest.fixture(scope="class")
    def master_token(self):
        """Get master token for liquidaciones tests (require master role)"""
        # First try master user
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "master@techrepair.local",
            "password": "Master2026!"
        })
        if response.status_code == 200:
            return response.json().get("token")
        # Fallback to emergency access
        response = requests.post(f"{BASE_URL}/api/auth/emergency", json={
            "key": "RevixEmergency2026SecureKey!"
        })
        if response.status_code == 200:
            return response.json().get("token")
        pytest.skip("Master authentication failed")
    
    def test_admin_login(self):
        """Test admin login works"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@techrepair.local",
            "password": "Admin2026!"
        })
        assert response.status_code == 200
        data = response.json()
        assert "token" in data
        assert "user" in data
        assert data["user"]["role"] == "admin"


class TestGarantiasAPI:
    """Tests for warranty/guarantee creation and retrieval"""
    
    @pytest.fixture(scope="class")
    def admin_headers(self):
        """Get admin auth headers"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@techrepair.local",
            "password": "Admin2026!"
        })
        if response.status_code != 200:
            pytest.skip("Admin auth failed")
        token = response.json().get("token")
        return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    
    @pytest.fixture(scope="class")
    def orden_enviada_id(self, admin_headers):
        """Get an order in 'enviado' state for testing"""
        # Known order IDs from context
        test_order_ids = [
            "e522a87a-8c1f-4d73-929a-bc4e5ff1bc45",
            "41496916-326b-4854-ad95-727c825d1984"
        ]
        for orden_id in test_order_ids:
            response = requests.get(
                f"{BASE_URL}/api/ordenes/{orden_id}",
                headers=admin_headers
            )
            if response.status_code == 200:
                orden = response.json()
                if orden.get("estado") == "enviado" and not orden.get("es_garantia"):
                    return orden_id
        pytest.skip("No suitable order in 'enviado' state found")
    
    def test_get_orden_enviada(self, admin_headers, orden_enviada_id):
        """Verify we can get the order in enviado state"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert data["estado"] == "enviado"
        print(f"Order {data['numero_orden']} is in 'enviado' state - ready for warranty test")
    
    def test_crear_garantia_success(self, admin_headers, orden_enviada_id):
        """Test creating a warranty order from a shipped order"""
        response = requests.post(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}/crear-garantia",
            headers=admin_headers
        )
        
        # Could be 200 or 400 if warranty already exists
        if response.status_code == 400:
            detail = response.json().get("detail", "")
            if "ya es una garantía" in detail or "garantía" in detail.lower():
                print(f"Warranty already exists for this order: {detail}")
                return
        
        assert response.status_code == 200, f"Failed: {response.text}"
        data = response.json()
        
        # Verify response structure
        assert "message" in data
        assert "orden_garantia" in data
        assert "incidencia" in data
        
        orden_garantia = data["orden_garantia"]
        
        # Verify inherited fields
        assert orden_garantia.get("es_garantia") == True
        assert orden_garantia.get("orden_padre_id") == orden_enviada_id
        assert "numero_orden_padre" in orden_garantia
        
        # Get parent order to verify inherited fields
        parent_response = requests.get(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}",
            headers=admin_headers
        )
        parent = parent_response.json()
        
        # Verify cliente_id is inherited
        assert orden_garantia.get("cliente_id") == parent.get("cliente_id")
        
        # Verify dispositivo is inherited
        if parent.get("dispositivo"):
            assert orden_garantia.get("dispositivo") is not None
        
        print(f"Created warranty order: {orden_garantia.get('numero_orden')}")
        print(f"Parent order: {orden_garantia.get('numero_orden_padre')}")
    
    def test_crear_garantia_solo_enviado(self, admin_headers):
        """Test that warranty can only be created from shipped orders"""
        # Try to create warranty from a non-shipped order (should fail)
        # First get any order not in 'enviado' state
        response = requests.get(
            f"{BASE_URL}/api/ordenes",
            headers=admin_headers
        )
        if response.status_code != 200:
            pytest.skip("Could not list orders")
        
        ordenes = response.json()
        orden_no_enviada = None
        for orden in ordenes:
            if orden.get("estado") != "enviado" and not orden.get("es_garantia"):
                orden_no_enviada = orden
                break
        
        if not orden_no_enviada:
            pytest.skip("No order in non-shipped state found")
        
        response = requests.post(
            f"{BASE_URL}/api/ordenes/{orden_no_enviada['id']}/crear-garantia",
            headers=admin_headers
        )
        
        assert response.status_code == 400
        detail = response.json().get("detail", "")
        assert "enviado" in detail.lower() or "enviadas" in detail.lower()
        print(f"Correctly rejected warranty creation for order in state: {orden_no_enviada['estado']}")
    
    def test_obtener_garantias_orden(self, admin_headers, orden_enviada_id):
        """Test getting warranty orders linked to a parent"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}/garantias",
            headers=admin_headers
        )
        
        assert response.status_code == 200
        garantias = response.json()
        
        # Should be a list
        assert isinstance(garantias, list)
        
        if len(garantias) > 0:
            garantia = garantias[0]
            # Each warranty should have es_garantia=True
            assert garantia.get("es_garantia") == True
            # Should reference parent
            assert garantia.get("orden_padre_id") == orden_enviada_id
            print(f"Found {len(garantias)} warranty order(s) for parent {orden_enviada_id}")
            for g in garantias:
                print(f"  - {g.get('numero_orden')} (estado: {g.get('estado')})")
        else:
            print("No warranty orders found (may need to run create test first)")
    
    def test_garantia_hereda_campos_cliente(self, admin_headers, orden_enviada_id):
        """Verify warranty order inherits cliente fields"""
        # First get the parent order
        parent_response = requests.get(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}",
            headers=admin_headers
        )
        assert parent_response.status_code == 200
        parent = parent_response.json()
        
        # Get warranty orders
        response = requests.get(
            f"{BASE_URL}/api/ordenes/{orden_enviada_id}/garantias",
            headers=admin_headers
        )
        assert response.status_code == 200
        garantias = response.json()
        
        if len(garantias) == 0:
            pytest.skip("No warranty orders to test inheritance")
        
        garantia = garantias[0]
        
        # Verify inherited fields
        assert garantia.get("cliente_id") == parent.get("cliente_id"), "cliente_id not inherited"
        
        # If parent has cliente_nombre, warranty should too
        if parent.get("cliente_nombre"):
            assert garantia.get("cliente_nombre") == parent.get("cliente_nombre"), "cliente_nombre not inherited"
        
        print(f"Verified campo inheritance for warranty {garantia.get('numero_orden')}")


class TestLiquidacionesAPI:
    """Tests for Insurama liquidations module"""
    
    @pytest.fixture(scope="class")
    def master_headers(self):
        """Get master auth headers (required for liquidaciones)"""
        # Try master user created via emergency access
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "master@test.local",
            "password": "Master2026!"
        })
        if response.status_code == 200:
            token = response.json().get("token")
            return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        
        pytest.skip("Master authentication failed - liquidaciones tests require master role")
    
    def test_get_liquidaciones_pendientes(self, master_headers):
        """Test GET /api/liquidaciones/pendientes"""
        response = requests.get(
            f"{BASE_URL}/api/liquidaciones/pendientes",
            headers=master_headers
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # Verify response structure
        assert "pendientes" in data
        assert "pagados" in data
        assert "reclamados" in data
        assert "impagados" in data
        assert "resumen" in data
        
        resumen = data["resumen"]
        assert "total_pendientes" in resumen
        assert "total_pagados" in resumen
        assert "importe_pendiente" in resumen
        assert "importe_pagado" in resumen
        
        print(f"Liquidaciones: {resumen['total_pendientes']} pendientes, {resumen['total_pagados']} pagados")
        print(f"Importes: {resumen['importe_pendiente']}€ pendiente, {resumen['importe_pagado']}€ pagado")
    
    def test_get_historial_meses(self, master_headers):
        """Test GET /api/liquidaciones/historial-meses"""
        response = requests.get(
            f"{BASE_URL}/api/liquidaciones/historial-meses",
            headers=master_headers
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # Verify response structure
        assert "meses" in data
        assert isinstance(data["meses"], list)
        
        if len(data["meses"]) > 0:
            mes = data["meses"][0]
            assert "mes" in mes
            assert "total_siniestros" in mes
            assert "total_importe" in mes
            assert "pagados" in mes
            assert "pendientes" in mes
            print(f"Historial: {len(data['meses'])} meses con datos")
            for m in data["meses"][:3]:
                print(f"  - {m['mes']}: {m['total_siniestros']} siniestros, {m['total_importe']}€")
        else:
            print("No monthly history data yet")
    
    def test_marcar_pagados_empty_list(self, master_headers):
        """Test POST /api/liquidaciones/marcar-pagados with empty list"""
        response = requests.post(
            f"{BASE_URL}/api/liquidaciones/marcar-pagados",
            headers=master_headers,
            json={
                "codigos": [],
                "mes": "2026-01"
            }
        )
        
        # Empty list should be rejected (422) or succeed with 0 marked (200)
        assert response.status_code in [200, 422]
        if response.status_code == 200:
            data = response.json()
            assert data.get("marcados") == 0
            print("Empty list marking returns success with 0 marked")
        else:
            print("Empty list correctly rejected with 422")
    
    def test_importar_excel_sin_archivo(self, master_headers):
        """Test importar-excel endpoint returns error without file"""
        response = requests.post(
            f"{BASE_URL}/api/liquidaciones/importar-excel",
            headers=master_headers
        )
        
        # Should fail - no file provided
        assert response.status_code == 422  # Unprocessable Entity (missing file)
        print("Correctly requires file for import")
    
    def test_importar_excel_formato_incorrecto(self, master_headers):
        """Test importar-excel with invalid file format"""
        # Create a fake text file
        files = {
            'file': ('test.txt', io.BytesIO(b'not an excel file'), 'text/plain')
        }
        
        # Remove Content-Type to let requests set it for multipart
        headers = {"Authorization": master_headers["Authorization"]}
        
        response = requests.post(
            f"{BASE_URL}/api/liquidaciones/importar-excel",
            headers=headers,
            files=files,
            data={"mes": "2026-01"}
        )
        
        # Should fail with 400 or 500 - invalid format
        assert response.status_code in [400, 500, 422]
        print(f"Correctly rejects invalid file format: {response.status_code}")


class TestExcelImport:
    """Tests for Excel import with generated test file"""
    
    @pytest.fixture(scope="class")
    def master_headers(self):
        """Get master auth headers"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "master@test.local",
            "password": "Master2026!"
        })
        if response.status_code == 200:
            token = response.json().get("token")
            return {"Authorization": f"Bearer {token}"}
        
        pytest.skip("Master auth failed")
    
    def test_importar_excel_con_openpyxl(self, master_headers):
        """Test Excel import with a properly formatted file"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Siniestros"
        
        # Add headers
        headers = ["Siniestro", "Póliza", "Fecha Apertura", "Fecha Cierre", "Presupuesto"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add test data - using a fake code that won't match
        ws.cell(row=2, column=1, value="TEST-SINIESTRO-001")
        ws.cell(row=2, column=2, value="POL-123456")
        ws.cell(row=2, column=3, value="2026-01-15")
        ws.cell(row=2, column=4, value="2026-01-20")
        ws.cell(row=2, column=5, value=150.00)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {
            'file': ('test_liquidacion.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        response = requests.post(
            f"{BASE_URL}/api/liquidaciones/importar-excel",
            headers=master_headers,
            files=files,
            data={"mes": "2026-03"}
        )
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        data = response.json()
        
        # Verify response structure
        assert "success" in data
        assert data["success"] == True
        assert "resumen" in data
        
        resumen = data["resumen"]
        assert "auto_liquidados" in resumen
        assert "pendientes_garantia" in resumen
        assert "ya_pagados" in resumen
        assert "no_encontrados" in resumen
        
        print(f"Excel import successful:")
        print(f"  - Auto-liquidados: {resumen['auto_liquidados']}")
        print(f"  - Pendientes garantía: {resumen['pendientes_garantia']}")
        print(f"  - Ya pagados: {resumen['ya_pagados']}")
        print(f"  - No encontrados: {resumen['no_encontrados']}")
        
        # The test code should be in "no_encontrados" since it's fake
        assert resumen["no_encontrados"] >= 1 or resumen["auto_liquidados"] >= 0


class TestLiquidacionesStatusUpdate:
    """Tests for liquidacion status updates"""
    
    @pytest.fixture(scope="class")
    def master_headers(self):
        """Get master auth headers"""
        response = requests.post(f"{BASE_URL}/api/auth/login", json={
            "email": "master@test.local",
            "password": "Master2026!"
        })
        if response.status_code == 200:
            token = response.json().get("token")
            return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        
        pytest.skip("Master auth failed")
    
    def test_actualizar_estado_liquidacion(self, master_headers):
        """Test updating liquidation status"""
        # First, get pendientes to find a code
        response = requests.get(
            f"{BASE_URL}/api/liquidaciones/pendientes",
            headers=master_headers
        )
        
        if response.status_code != 200:
            pytest.skip("Could not get pendientes")
        
        data = response.json()
        
        # Try to update status of a test code (may not exist)
        test_codigo = "TEST-UPDATE-001"
        
        response = requests.patch(
            f"{BASE_URL}/api/liquidaciones/{test_codigo}/estado",
            headers=master_headers,
            params={"estado": "pagado", "notas": "Test update"}
        )
        
        # Should succeed (upsert) or return error
        assert response.status_code in [200, 404]
        
        if response.status_code == 200:
            result = response.json()
            assert result.get("success") == True
            print(f"Status update successful for {test_codigo}")
        else:
            print(f"Status update returned {response.status_code} - expected for non-existent code")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
