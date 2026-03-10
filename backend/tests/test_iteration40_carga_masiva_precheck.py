"""
Iteration 40: Testing Carga Masiva Precheck Flow for Insurama
Tests the new pre-check endpoint and confirmation flow for bulk loading claims
"""
import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestInsuramaPrecheck:
    """Test suite for /api/insurama/carga-masiva/precheck endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup: Get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login to get token
        login_resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "ramiraz91@gmail.com",
            "password": "temp123"
        })
        
        if login_resp.status_code != 200:
            pytest.skip("Could not authenticate - skipping authenticated tests")
        
        # Note: API returns 'token' not 'access_token'
        self.token = login_resp.json().get("token")
        self.session.headers.update({"Authorization": f"Bearer {self.token}"})
    
    def test_precheck_returns_400_for_non_excel_file(self):
        """POST /api/insurama/carga-masiva/precheck returns 400 for non-Excel file"""
        # Create a fake text file
        fake_file = BytesIO(b"This is not an excel file")
        
        # Use token for auth in multipart request
        files = {"file": ("test.txt", fake_file, "text/plain")}
        headers = {"Authorization": f"Bearer {self.token}"}
        
        resp = requests.post(
            f"{BASE_URL}/api/insurama/carga-masiva/precheck",
            headers=headers,
            files=files
        )
        
        assert resp.status_code == 400, f"Expected 400, got {resp.status_code}: {resp.text}"
        data = resp.json()
        assert "detail" in data
        assert "Excel" in data["detail"] or "xlsx" in data["detail"].lower() or "xls" in data["detail"].lower()
        print(f"✓ Non-Excel file returns 400 with message: {data['detail']}")
    
    def test_precheck_returns_400_for_csv_file(self):
        """POST /api/insurama/carga-masiva/precheck returns 400 for CSV file"""
        fake_csv = BytesIO(b"col1,col2\nval1,val2")
        
        headers = {"Authorization": f"Bearer {self.token}"}
        files = {"file": ("test.csv", fake_csv, "text/csv")}
        
        resp = requests.post(
            f"{BASE_URL}/api/insurama/carga-masiva/precheck",
            headers=headers,
            files=files
        )
        
        assert resp.status_code == 400, f"Expected 400, got {resp.status_code}: {resp.text}"
        print(f"✓ CSV file returns 400")
    
    def test_precheck_endpoint_exists_and_requires_auth(self):
        """POST /api/insurama/carga-masiva/precheck endpoint exists (not 404)"""
        # Request without auth
        fake_excel = BytesIO(b"fake excel content")
        files = {"file": ("test.xlsx", fake_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        resp = requests.post(
            f"{BASE_URL}/api/insurama/carga-masiva/precheck",
            files=files
        )
        
        # Should be 401 (unauthorized) not 404 (not found)
        assert resp.status_code in [401, 403], f"Expected 401/403 without auth, got {resp.status_code}"
        print(f"✓ Endpoint exists and requires auth (status: {resp.status_code})")
    
    def test_carga_masiva_endpoint_exists_and_requires_auth(self):
        """POST /api/insurama/carga-masiva endpoint exists (not 404)"""
        fake_excel = BytesIO(b"fake excel content")
        files = {"file": ("test.xlsx", fake_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        resp = requests.post(
            f"{BASE_URL}/api/insurama/carga-masiva",
            files=files
        )
        
        # Should be 401 (unauthorized) not 404 (not found)
        assert resp.status_code in [401, 403], f"Expected 401/403 without auth, got {resp.status_code}"
        print(f"✓ carga-masiva endpoint exists and requires auth (status: {resp.status_code})")
    
    def test_precheck_with_valid_excel_structure_returns_summary(self):
        """POST /api/insurama/carga-masiva/precheck with valid Excel returns summary structure"""
        # Create a minimal valid Excel file using openpyxl
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            # Use column name that the backend expects
            ws['A1'] = 'Numero de siniestro'
            ws['A2'] = 'TEST123456'
            ws['A3'] = 'TEST789012'
            ws['A4'] = ''  # Empty row
            ws['A5'] = 'TEST123456'  # Duplicate
            ws['A6'] = 'AB'  # Invalid format (too short)
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            headers = {"Authorization": f"Bearer {self.token}"}
            files = {"file": ("test_siniestros.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            
            resp = requests.post(
                f"{BASE_URL}/api/insurama/carga-masiva/precheck",
                headers=headers,
                files=files
            )
            
            assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
            data = resp.json()
            
            # Verify response structure
            assert "archivo" in data, "Response should contain 'archivo'"
            assert "columna_codigo" in data, "Response should contain 'columna_codigo'"
            assert "resumen" in data, "Response should contain 'resumen'"
            assert "detalles" in data, "Response should contain 'detalles'"
            
            resumen = data["resumen"]
            assert "total_filas" in resumen, "resumen should contain 'total_filas'"
            assert "validos_unicos" in resumen, "resumen should contain 'validos_unicos'"
            assert "vacios" in resumen, "resumen should contain 'vacios'"
            assert "duplicados" in resumen, "resumen should contain 'duplicados'"
            assert "formato_invalido" in resumen, "resumen should contain 'formato_invalido'"
            assert "existentes_en_crm" in resumen, "resumen should contain 'existentes_en_crm'"
            assert "nuevos_estimados" in resumen, "resumen should contain 'nuevos_estimados'"
            assert "listos_para_procesar" in resumen, "resumen should contain 'listos_para_procesar'"
            
            print(f"✓ Precheck returns valid summary structure")
            print(f"  - Total filas: {resumen['total_filas']}")
            print(f"  - Válidos únicos: {resumen['validos_unicos']}")
            print(f"  - Vacíos: {resumen['vacios']}")
            print(f"  - Duplicados: {resumen['duplicados']}")
            print(f"  - Formato inválido: {resumen['formato_invalido']}")
            print(f"  - Existentes CRM: {resumen['existentes_en_crm']}")
            print(f"  - Nuevos estimados: {resumen['nuevos_estimados']}")
            print(f"  - Listos para procesar: {resumen['listos_para_procesar']}")
            
            # Verify counts match expectations based on test data
            assert resumen['total_filas'] == 5, f"Expected 5 rows, got {resumen['total_filas']}"
            assert resumen['vacios'] >= 1, f"Expected at least 1 empty, got {resumen['vacios']}"
            assert resumen['duplicados'] >= 1, f"Expected at least 1 duplicate, got {resumen['duplicados']}"
            assert resumen['formato_invalido'] >= 1, f"Expected at least 1 invalid format, got {resumen['formato_invalido']}"
            
        except ImportError:
            pytest.skip("openpyxl not installed - cannot create test Excel file")
    
    def test_precheck_detalles_contain_correct_status_values(self):
        """POST /api/insurama/carga-masiva/precheck detalles contain proper status values"""
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'codigo_siniestro'
            ws['A2'] = 'VALID123456'
            ws['A3'] = ''  # Empty
            ws['A4'] = 'AB'  # Invalid format
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            headers = {"Authorization": f"Bearer {self.token}"}
            files = {"file": ("test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            
            resp = requests.post(
                f"{BASE_URL}/api/insurama/carga-masiva/precheck",
                headers=headers,
                files=files
            )
            
            assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
            data = resp.json()
            detalles = data.get("detalles", [])
            
            # Check that detalles have proper structure
            for detalle in detalles:
                assert "fila" in detalle, "detalle should have 'fila'"
                assert "codigo" in detalle, "detalle should have 'codigo'"
                assert "status" in detalle, "detalle should have 'status'"
                assert "mensaje" in detalle, "detalle should have 'mensaje'"
                
                # Status should be one of: nuevo, existente, vacio, duplicado, formato_invalido
                valid_statuses = ["nuevo", "existente", "vacio", "duplicado", "formato_invalido", "listo"]
                assert detalle["status"] in valid_statuses, f"Invalid status: {detalle['status']}"
            
            print(f"✓ Detalles contain valid status values")
            print(f"  - Total detalles: {len(detalles)}")
            
        except ImportError:
            pytest.skip("openpyxl not installed")


class TestInsuramaEndpointRegression:
    """Regression tests for Insurama endpoints"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup: Get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        login_resp = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "ramiraz91@gmail.com",
            "password": "temp123"
        })
        
        if login_resp.status_code != 200:
            pytest.skip("Could not authenticate")
        
        # Note: API returns 'token' not 'access_token'
        token = login_resp.json().get("token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
    
    def test_insurama_config_endpoint(self):
        """GET /api/insurama/config returns 200"""
        resp = self.session.get(f"{BASE_URL}/api/insurama/config")
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
        data = resp.json()
        assert "configurado" in data
        print(f"✓ GET /api/insurama/config returns 200 (configurado: {data['configurado']})")
    
    def test_insurama_presupuestos_endpoint(self):
        """GET /api/insurama/presupuestos returns 200 (if configured)"""
        # First check if configured
        config_resp = self.session.get(f"{BASE_URL}/api/insurama/config")
        config = config_resp.json()
        
        if not config.get("configurado") or not config.get("conexion_ok"):
            pytest.skip("Sumbroker not configured - skipping presupuestos test")
        
        resp = self.session.get(f"{BASE_URL}/api/insurama/presupuestos?limit=5")
        # Can be 200 or 400 depending on credentials
        assert resp.status_code in [200, 400], f"Expected 200 or 400, got {resp.status_code}"
        print(f"✓ GET /api/insurama/presupuestos endpoint accessible (status: {resp.status_code})")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
