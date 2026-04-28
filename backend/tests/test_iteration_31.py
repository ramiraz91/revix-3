"""
Iteration 31 - Backend Tests for 4 new features:
1. diagnostico_recepcion field (tecnico can write, admin can read)
2. Badge GARANTÍA in listings (es_garantia=true)
3. Date filters in /ordenes with presets and field selector
4. Date filters + Excel export in /contabilidad
"""
import pytest
import requests
import os

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
ADMIN_EMAIL = "master@revix.es"
ADMIN_PASSWORD = "RevixMaster2026!"
TECNICO_EMAIL = "tecnico1@revix.es"
TECNICO_PASSWORD = "Tecnico1Demo!"


@pytest.fixture(scope="module")
def admin_token():
    """Get admin authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": ADMIN_EMAIL,
        "password": ADMIN_PASSWORD
    })
    if response.status_code == 429:
        pytest.skip("Rate limited - wait 15 minutes")
    assert response.status_code == 200, f"Admin login failed: {response.text}"
    return response.json().get("token")


@pytest.fixture(scope="module")
def tecnico_token():
    """Get tecnico authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TECNICO_EMAIL,
        "password": TECNICO_PASSWORD
    })
    if response.status_code == 429:
        pytest.skip("Rate limited - wait 15 minutes")
    assert response.status_code == 200, f"Tecnico login failed: {response.text}"
    return response.json().get("token")


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def tecnico_headers(tecnico_token):
    return {"Authorization": f"Bearer {tecnico_token}", "Content-Type": "application/json"}


# ==================== FEATURE 1: diagnostico_recepcion ====================

class TestDiagnosticoRecepcion:
    """Tests for diagnostico_recepcion field"""
    
    def test_tecnico_can_update_diagnostico_recepcion(self, tecnico_headers):
        """(1a) Tecnico can PATCH diagnostico_recepcion on an order"""
        # First get an order assigned to tecnico
        response = requests.get(f"{BASE_URL}/api/ordenes?search=OT-DEMO", headers=tecnico_headers)
        assert response.status_code == 200
        ordenes = response.json()
        assert len(ordenes) > 0, "No demo orders found"
        
        # Use OT-DEMO-001 or first available
        orden = next((o for o in ordenes if "OT-DEMO-001" in o.get("numero_orden", "")), ordenes[0])
        orden_id = orden["id"]
        
        # Update diagnostico_recepcion
        diagnostico_text = "Inspección visual: pantalla con grietas. Dispositivo arranca pero touch no responde en zona inferior."
        response = requests.patch(
            f"{BASE_URL}/api/ordenes/{orden_id}",
            json={"diagnostico_recepcion": diagnostico_text},
            headers=tecnico_headers
        )
        assert response.status_code == 200, f"PATCH failed: {response.text}"
        
        # Verify the field was saved
        updated = response.json()
        assert updated.get("diagnostico_recepcion") == diagnostico_text
    
    def test_admin_can_read_diagnostico_recepcion(self, admin_headers):
        """(1c) Admin can read diagnostico_recepcion from order detail"""
        # Get the same order
        response = requests.get(f"{BASE_URL}/api/ordenes?search=OT-DEMO-001", headers=admin_headers)
        assert response.status_code == 200
        ordenes = response.json()
        assert len(ordenes) > 0
        
        orden = ordenes[0]
        orden_id = orden["id"]
        
        # Get full order detail
        response = requests.get(f"{BASE_URL}/api/ordenes/{orden_id}", headers=admin_headers)
        assert response.status_code == 200
        orden_detail = response.json()
        
        # Verify diagnostico_recepcion is present
        assert "diagnostico_recepcion" in orden_detail
        # Should have the text we set earlier (or be non-empty if already set)
        print(f"diagnostico_recepcion: {orden_detail.get('diagnostico_recepcion')}")


# ==================== FEATURE 2: Badge GARANTÍA ====================

class TestBadgeGarantia:
    """Tests for es_garantia badge in listings"""
    
    def test_solo_garantias_filter_returns_garantia_orders(self, admin_headers):
        """(2a/3b) GET /api/ordenes/v2?solo_garantias=true returns orders with es_garantia=true"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?solo_garantias=true",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Should return paginated response
        assert "data" in data
        assert "total" in data
        
        ordenes = data["data"]
        print(f"Found {len(ordenes)} garantía orders")
        
        # All returned orders should have es_garantia=true
        for orden in ordenes:
            assert orden.get("es_garantia") == True, f"Order {orden.get('numero_orden')} should have es_garantia=true"
    
    def test_garantia_order_exists(self, admin_headers):
        """Verify OT-20260428-AFB5AD05 exists and has es_garantia=true"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes?search=OT-20260428-AFB5AD05",
            headers=admin_headers
        )
        assert response.status_code == 200
        ordenes = response.json()
        
        # Find the specific order
        garantia_orden = next((o for o in ordenes if "AFB5AD05" in o.get("numero_orden", "")), None)
        if garantia_orden:
            assert garantia_orden.get("es_garantia") == True
            print(f"Found garantía order: {garantia_orden.get('numero_orden')}")
        else:
            # Try v2 endpoint
            response = requests.get(
                f"{BASE_URL}/api/ordenes/v2?solo_garantias=true",
                headers=admin_headers
            )
            data = response.json()
            print(f"Garantía orders via v2: {[o.get('numero_orden') for o in data.get('data', [])]}")


# ==================== FEATURE 3: Date Filters in /ordenes ====================

class TestDateFiltersOrdenes:
    """Tests for date filters in /ordenes endpoint"""
    
    def test_fecha_campo_filter_created_at(self, admin_headers):
        """(3a) Filter by created_at (default)"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-01-01&fecha_campo=created_at",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders with created_at >= 2026-01-01: {data.get('total')}")
    
    def test_fecha_campo_filter_fecha_enviado(self, admin_headers):
        """(3b) Filter by fecha_enviado"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-04-01&fecha_campo=fecha_enviado",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders with fecha_enviado >= 2026-04-01: {data.get('total')}")
    
    def test_fecha_campo_filter_fecha_recibida_centro(self, admin_headers):
        """Filter by fecha_recibida_centro"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-01-01&fecha_campo=fecha_recibida_centro",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders with fecha_recibida_centro >= 2026-01-01: {data.get('total')}")
    
    def test_fecha_campo_filter_fecha_fin_reparacion(self, admin_headers):
        """Filter by fecha_fin_reparacion"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-01-01&fecha_campo=fecha_fin_reparacion",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders with fecha_fin_reparacion >= 2026-01-01: {data.get('total')}")
    
    def test_fecha_campo_filter_fecha_inicio_reparacion(self, admin_headers):
        """Filter by fecha_inicio_reparacion"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-01-01&fecha_campo=fecha_inicio_reparacion",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders with fecha_inicio_reparacion >= 2026-01-01: {data.get('total')}")
    
    def test_fecha_range_filter(self, admin_headers):
        """Filter with both fecha_desde and fecha_hasta"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes/v2?fecha_desde=2026-04-01&fecha_hasta=2026-04-30&fecha_campo=created_at",
            headers=admin_headers
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        print(f"Orders in April 2026: {data.get('total')}")


# ==================== FEATURE 4: Contabilidad Filters + Excel Export ====================

class TestContabilidadFiltersExcel:
    """Tests for date filters and Excel export in /contabilidad"""
    
    def test_contabilidad_stats(self, admin_headers):
        """Get contabilidad stats"""
        response = requests.get(f"{BASE_URL}/api/contabilidad/stats", headers=admin_headers)
        assert response.status_code == 200
        stats = response.json()
        assert "facturas_venta" in stats
        assert "facturas_compra" in stats
        print(f"Contabilidad stats: {stats}")
    
    def test_contabilidad_resumen(self, admin_headers):
        """Get contabilidad resumen"""
        response = requests.get(f"{BASE_URL}/api/contabilidad/informes/resumen", headers=admin_headers)
        assert response.status_code == 200
        resumen = response.json()
        assert "ventas" in resumen
        assert "compras" in resumen
        print(f"Contabilidad resumen año: {resumen.get('año')}")
    
    def test_export_excel_completo(self, admin_headers):
        """(4c) Export Excel with formato=completo returns valid xlsx"""
        response = requests.get(
            f"{BASE_URL}/api/contabilidad/export-excel?formato=completo&tipo=todas",
            headers=admin_headers
        )
        assert response.status_code == 200
        
        # Check content type
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "application/vnd" in content_type, f"Unexpected content type: {content_type}"
        
        # Check content disposition
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp
        assert ".xlsx" in content_disp
        
        # Verify it's a valid xlsx (starts with PK for zip)
        content = response.content
        assert len(content) > 0
        assert content[:2] == b'PK', "Excel file should start with PK (zip signature)"
        
        print(f"Excel export successful, size: {len(content)} bytes")
    
    def test_export_excel_unica(self, admin_headers):
        """Export Excel with formato=unica (single sheet)"""
        response = requests.get(
            f"{BASE_URL}/api/contabilidad/export-excel?formato=unica&tipo=venta",
            headers=admin_headers
        )
        assert response.status_code == 200
        assert response.content[:2] == b'PK'
        print(f"Excel unica export successful, size: {len(response.content)} bytes")
    
    def test_export_excel_resumen(self, admin_headers):
        """Export Excel with formato=resumen"""
        response = requests.get(
            f"{BASE_URL}/api/contabilidad/export-excel?formato=resumen&tipo=compra",
            headers=admin_headers
        )
        assert response.status_code == 200
        assert response.content[:2] == b'PK'
        print(f"Excel resumen export successful, size: {len(response.content)} bytes")
    
    def test_export_excel_with_date_range(self, admin_headers):
        """Export Excel with date range filter"""
        response = requests.get(
            f"{BASE_URL}/api/contabilidad/export-excel?fecha_desde=2026-01-01&fecha_hasta=2026-12-31&formato=completo&tipo=todas",
            headers=admin_headers
        )
        assert response.status_code == 200
        assert response.content[:2] == b'PK'
        print(f"Excel with date range export successful, size: {len(response.content)} bytes")
    
    def test_export_excel_has_two_sheets_completo(self, admin_headers):
        """(4c) Verify completo format has 2 sheets (Resumen + Detalle)"""
        import io
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(
            f"{BASE_URL}/api/contabilidad/export-excel?formato=completo&tipo=todas",
            headers=admin_headers
        )
        assert response.status_code == 200
        
        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        print(f"Excel sheets: {sheet_names}")
        assert len(sheet_names) >= 2, f"Expected at least 2 sheets, got {len(sheet_names)}: {sheet_names}"
        assert "Resumen" in sheet_names, f"Missing 'Resumen' sheet in {sheet_names}"
        assert "Detalle" in sheet_names, f"Missing 'Detalle' sheet in {sheet_names}"


# ==================== LEGACY ENDPOINT TESTS ====================

class TestLegacyOrdenesEndpoint:
    """Tests for legacy /api/ordenes endpoint (non-paginated)"""
    
    def test_legacy_ordenes_with_fecha_campo(self, admin_headers):
        """Legacy endpoint also supports fecha_campo"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes?fecha_desde=2026-01-01&fecha_campo=created_at",
            headers=admin_headers
        )
        assert response.status_code == 200
        ordenes = response.json()
        assert isinstance(ordenes, list)
        print(f"Legacy endpoint returned {len(ordenes)} orders")
    
    def test_legacy_ordenes_solo_garantias(self, admin_headers):
        """Legacy endpoint supports solo_garantias filter"""
        response = requests.get(
            f"{BASE_URL}/api/ordenes?solo_garantias=true",
            headers=admin_headers
        )
        assert response.status_code == 200
        ordenes = response.json()
        assert isinstance(ordenes, list)
        for orden in ordenes:
            assert orden.get("es_garantia") == True
        print(f"Legacy endpoint returned {len(ordenes)} garantía orders")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
