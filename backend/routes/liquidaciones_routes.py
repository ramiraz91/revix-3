"""
Rutas para el Sistema de Liquidaciones de Insurama
Gestión de pagos mensuales de siniestros reparados
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
import uuid
import io

from config import db, logger
from auth import require_master

router = APIRouter(prefix="/liquidaciones", tags=["liquidaciones"])


# ==================== MODELOS ====================

class LiquidacionItem(BaseModel):
    codigo_siniestro: str
    poliza: Optional[str] = None
    fecha_apertura: Optional[str] = None
    fecha_cierre: Optional[str] = None
    tipo_reparacion: Optional[str] = None
    tipologia_siniestro: Optional[str] = None
    importe_presupuesto: float = 0
    compania: Optional[str] = None
    producto: Optional[str] = None
    
    # Estados de liquidación
    estado: str = "pendiente"  # pendiente, pagado, reclamado, en_resolucion
    fecha_pago: Optional[str] = None
    
    # Garantías
    tiene_garantia: bool = False
    codigo_garantia: Optional[str] = None
    costes_garantia: float = 0
    
    # Metadata
    mes_liquidacion: str  # Formato: "2026-01"
    notas: Optional[str] = None


class ImportarLiquidacionRequest(BaseModel):
    mes: str  # Formato: "2026-01"


# ==================== ENDPOINTS ====================

@router.get("/pendientes")
async def obtener_liquidaciones_pendientes(user: dict = Depends(require_master)):
    """
    Obtiene los siniestros enviados (validados por admin) que están pendientes de liquidar.
    Estos son las órdenes con estado 'enviado' que tienen código de Insurama.
    """
    try:
        # Buscar órdenes con estado "enviado" que tengan código de Insurama
        pipeline = [
            {
                "$match": {
                    "estado": "enviado",
                    "codigo_insurama": {"$exists": True, "$ne": None, "$ne": ""}
                }
            },
            {
                "$lookup": {
                    "from": "liquidaciones",
                    "localField": "codigo_insurama",
                    "foreignField": "codigo_siniestro",
                    "as": "liquidacion"
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "orden_id": "$id",
                    "numero_orden": 1,
                    "codigo_insurama": 1,
                    "cliente_nombre": 1,
                    "dispositivo_marca": 1,
                    "dispositivo_modelo": 1,
                    "total": 1,
                    "fecha_envio": 1,
                    "fecha_creacion": "$created_at",
                    "liquidacion": {"$arrayElemAt": ["$liquidacion", 0]}
                }
            },
            {"$sort": {"fecha_envio": -1}}
        ]
        
        ordenes = await db.ordenes.aggregate(pipeline).to_list(500)
        
        # Separar por estado de liquidación
        pendientes = []
        pagados = []
        reclamados = []
        
        for orden in ordenes:
            liq = orden.get("liquidacion")
            item = {
                "orden_id": orden.get("orden_id"),
                "numero_orden": orden.get("numero_orden"),
                "codigo_siniestro": orden.get("codigo_insurama"),
                "cliente": orden.get("cliente_nombre"),
                "dispositivo": f"{orden.get('dispositivo_marca', '')} {orden.get('dispositivo_modelo', '')}".strip(),
                "importe": orden.get("total", 0),
                "fecha_envio": orden.get("fecha_envio"),
                "estado_liquidacion": liq.get("estado") if liq else "pendiente",
                "fecha_pago": liq.get("fecha_pago") if liq else None,
                "tiene_garantia": liq.get("tiene_garantia", False) if liq else False,
                "codigo_garantia": liq.get("codigo_garantia") if liq else None,
                "costes_garantia": liq.get("costes_garantia", 0) if liq else 0,
                "mes_liquidacion": liq.get("mes_liquidacion") if liq else None,
                "notas": liq.get("notas") if liq else None
            }
            
            estado = item["estado_liquidacion"]
            if estado == "pagado":
                pagados.append(item)
            elif estado in ["reclamado", "en_resolucion"]:
                reclamados.append(item)
            else:
                pendientes.append(item)
        
        # Detectar impagados (más de 60 días desde envío)
        fecha_limite = (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()
        impagados = [p for p in pendientes if p.get("fecha_envio") and p["fecha_envio"] < fecha_limite]
        
        return {
            "pendientes": pendientes,
            "pagados": pagados,
            "reclamados": reclamados,
            "impagados": impagados,
            "resumen": {
                "total_pendientes": len(pendientes),
                "total_pagados": len(pagados),
                "total_reclamados": len(reclamados),
                "total_impagados": len(impagados),
                "importe_pendiente": sum(p.get("importe", 0) for p in pendientes),
                "importe_pagado": sum(p.get("importe", 0) for p in pagados),
                "importe_reclamado": sum(p.get("importe", 0) for p in reclamados)
            }
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo liquidaciones pendientes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/por-mes/{mes}")
async def obtener_liquidaciones_por_mes(mes: str, user: dict = Depends(require_master)):
    """
    Obtiene liquidaciones de un mes específico.
    Formato mes: "2026-01"
    """
    try:
        liquidaciones = await db.liquidaciones.find(
            {"mes_liquidacion": mes},
            {"_id": 0}
        ).sort("fecha_cierre", -1).to_list(500)
        
        # Calcular totales
        total_importe = sum(l.get("importe_presupuesto", 0) for l in liquidaciones)
        total_garantias = sum(l.get("costes_garantia", 0) for l in liquidaciones)
        pagados = [l for l in liquidaciones if l.get("estado") == "pagado"]
        pendientes = [l for l in liquidaciones if l.get("estado") == "pendiente"]
        reclamados = [l for l in liquidaciones if l.get("estado") in ["reclamado", "en_resolucion"]]
        
        return {
            "mes": mes,
            "liquidaciones": liquidaciones,
            "resumen": {
                "total_siniestros": len(liquidaciones),
                "total_importe": total_importe,
                "total_garantias": total_garantias,
                "importe_neto": total_importe - total_garantias,
                "pagados": len(pagados),
                "pendientes": len(pendientes),
                "reclamados": len(reclamados),
                "importe_pagado": sum(l.get("importe_presupuesto", 0) for l in pagados),
                "importe_pendiente": sum(l.get("importe_presupuesto", 0) for l in pendientes)
            }
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo liquidaciones del mes {mes}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/{codigo_siniestro}/estado")
async def actualizar_estado_liquidacion(
    codigo_siniestro: str,
    estado: str,
    notas: Optional[str] = None,
    user: dict = Depends(require_master)
):
    """Actualiza el estado de una liquidación"""
    try:
        estados_validos = ["pendiente", "pagado", "reclamado", "en_resolucion"]
        if estado not in estados_validos:
            raise HTTPException(status_code=400, detail=f"Estado inválido. Debe ser: {estados_validos}")
        
        update_data = {
            "estado": estado,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        if estado == "pagado":
            update_data["fecha_pago"] = datetime.now(timezone.utc).isoformat()
        
        if notas:
            update_data["notas"] = notas
        
        result = await db.liquidaciones.update_one(
            {"codigo_siniestro": codigo_siniestro},
            {"$set": update_data},
            upsert=True
        )
        
        return {"success": True, "codigo": codigo_siniestro, "nuevo_estado": estado}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error actualizando estado liquidación: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/{codigo_siniestro}/garantia")
async def actualizar_garantia_liquidacion(
    codigo_siniestro: str,
    tiene_garantia: bool,
    codigo_garantia: Optional[str] = None,
    costes_garantia: float = 0,
    user: dict = Depends(require_master)
):
    """Actualiza información de garantía de una liquidación"""
    try:
        update_data = {
            "tiene_garantia": tiene_garantia,
            "codigo_garantia": codigo_garantia,
            "costes_garantia": costes_garantia,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = await db.liquidaciones.update_one(
            {"codigo_siniestro": codigo_siniestro},
            {"$set": update_data},
            upsert=True
        )
        
        return {"success": True, "codigo": codigo_siniestro}
        
    except Exception as e:
        logger.error(f"Error actualizando garantía: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/importar-excel")
async def importar_liquidacion_excel(
    file: UploadFile = File(...),
    mes: str = None,
    user: dict = Depends(require_master)
):
    """
    Importa una liquidación desde un archivo Excel de Insurama.
    Auto-cruza los códigos con las órdenes del sistema:
    - Si la orden existe y no tiene garantía pendiente → marca como PAGADO
    - Si tiene garantía pendiente → marca como PENDIENTE (requiere revisión)
    - Duplicados (ya pagados) → se ignoran
    """
    try:
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        if "Siniestros" not in wb.sheetnames:
            # Intentar con la primera hoja
            ws = wb.active
            if not ws:
                raise HTTPException(status_code=400, detail="No se encontró la hoja 'Siniestros' en el archivo")
        else:
            ws = wb["Siniestros"]
        
        headers = [cell.value for cell in ws[1]]
        
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_lower = header.lower().strip()
                if "siniestro" in header_lower and "tipología" not in header_lower:
                    col_map["siniestro"] = idx
                elif "póliza" in header_lower or "poliza" in header_lower:
                    col_map["poliza"] = idx
                elif "fecha apertura" in header_lower:
                    col_map["fecha_apertura"] = idx
                elif "fecha cierre" in header_lower:
                    col_map["fecha_cierre"] = idx
                elif "tipo reparación" in header_lower or "tipo reparacion" in header_lower:
                    col_map["tipo_reparacion"] = idx
                elif "tipología" in header_lower or "tipologia" in header_lower:
                    col_map["tipologia"] = idx
                elif "presupuesto" in header_lower:
                    col_map["presupuesto"] = idx
                elif "compañía" in header_lower or "compania" in header_lower:
                    col_map["compania"] = idx
                elif "producto" in header_lower:
                    col_map["producto"] = idx
        
        if "siniestro" not in col_map:
            raise HTTPException(status_code=400, detail="No se encontró la columna 'Siniestro' en el archivo")
        
        if not mes:
            mes = datetime.now(timezone.utc).strftime("%Y-%m")
        
        now = datetime.now(timezone.utc)
        fecha_pago = now.isoformat()
        
        # Resultados detallados
        auto_liquidados = []       # Cruzados y pagados automáticamente
        pendientes_garantia = []   # Tienen garantía pendiente, necesitan revisión
        ya_pagados = []            # Duplicados: ya estaban pagados
        no_encontrados = []        # Sin orden en el sistema
        errores = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                codigo = row[col_map["siniestro"]] if col_map.get("siniestro") is not None else None
                if not codigo or not str(codigo).strip():
                    continue
                
                codigo = str(codigo).strip().upper()
                
                # Parsear importe
                importe = 0
                if col_map.get("presupuesto") is not None:
                    val = row[col_map["presupuesto"]]
                    if val:
                        try:
                            importe = float(str(val).replace(",", ".").replace("€", "").strip())
                        except:
                            pass
                
                # Parsear fechas
                fecha_apertura = None
                fecha_cierre = None
                if col_map.get("fecha_apertura") is not None:
                    val = row[col_map["fecha_apertura"]]
                    if val:
                        fecha_apertura = val.isoformat() if isinstance(val, datetime) else str(val)
                if col_map.get("fecha_cierre") is not None:
                    val = row[col_map["fecha_cierre"]]
                    if val:
                        fecha_cierre = val.isoformat() if isinstance(val, datetime) else str(val)
                
                liquidacion_data = {
                    "codigo_siniestro": codigo,
                    "poliza": str(row[col_map["poliza"]]) if col_map.get("poliza") is not None and row[col_map["poliza"]] else None,
                    "fecha_apertura": fecha_apertura,
                    "fecha_cierre": fecha_cierre,
                    "tipo_reparacion": str(row[col_map["tipo_reparacion"]]) if col_map.get("tipo_reparacion") is not None and row[col_map["tipo_reparacion"]] else None,
                    "tipologia_siniestro": str(row[col_map["tipologia"]]) if col_map.get("tipologia") is not None and row[col_map["tipologia"]] else None,
                    "importe_presupuesto": importe,
                    "compania": str(row[col_map["compania"]]) if col_map.get("compania") is not None and row[col_map["compania"]] else None,
                    "producto": str(row[col_map["producto"]]) if col_map.get("producto") is not None and row[col_map["producto"]] else None,
                    "mes_liquidacion": mes,
                    "updated_at": fecha_pago
                }
                
                # 1. Verificar si ya existe y está pagado (duplicado)
                existente = await db.liquidaciones.find_one({"codigo_siniestro": codigo})
                if existente and existente.get("estado") == "pagado":
                    ya_pagados.append({"codigo": codigo, "importe": importe, "fecha_pago": existente.get("fecha_pago")})
                    continue
                
                # 2. Buscar la orden en el sistema por código Insurama
                orden = await db.ordenes.find_one(
                    {"codigo_insurama": {"$regex": f"^{codigo}$", "$options": "i"}},
                    {"_id": 0, "id": 1, "numero_orden": 1, "ordenes_garantia": 1, "estado": 1, "cliente_nombre": 1}
                )
                
                if not orden:
                    # Sin orden en el sistema - registrar pero no marcar como pagado
                    liquidacion_data["estado"] = "pendiente"
                    liquidacion_data["nota_auto"] = "Código no encontrado en órdenes del sistema"
                    if existente:
                        await db.liquidaciones.update_one({"codigo_siniestro": codigo}, {"$set": liquidacion_data})
                    else:
                        liquidacion_data["created_at"] = fecha_pago
                        await db.liquidaciones.insert_one(liquidacion_data)
                    no_encontrados.append({"codigo": codigo, "importe": importe})
                    continue
                
                # 3. Verificar garantías pendientes
                garantia_ids = orden.get("ordenes_garantia", [])
                tiene_garantia_pendiente = False
                if garantia_ids:
                    estados_finales = ["enviado", "cancelado"]
                    garantias_pendientes = await db.ordenes.count_documents({
                        "id": {"$in": garantia_ids},
                        "estado": {"$nin": estados_finales}
                    })
                    tiene_garantia_pendiente = garantias_pendientes > 0
                
                if tiene_garantia_pendiente:
                    # Tiene garantía pendiente - registrar pero NO marcar como pagado
                    liquidacion_data["estado"] = "pendiente"
                    liquidacion_data["tiene_garantia"] = True
                    liquidacion_data["nota_auto"] = "Garantía pendiente de resolución"
                    liquidacion_data["orden_id"] = orden["id"]
                    if existente:
                        await db.liquidaciones.update_one({"codigo_siniestro": codigo}, {"$set": liquidacion_data})
                    else:
                        liquidacion_data["created_at"] = fecha_pago
                        await db.liquidaciones.insert_one(liquidacion_data)
                    pendientes_garantia.append({
                        "codigo": codigo, "importe": importe,
                        "numero_orden": orden.get("numero_orden"),
                        "motivo": "Garantía pendiente"
                    })
                    continue
                
                # 4. Todo OK → Marcar como PAGADO automáticamente
                liquidacion_data["estado"] = "pagado"
                liquidacion_data["fecha_pago"] = fecha_pago
                liquidacion_data["orden_id"] = orden["id"]
                liquidacion_data["auto_liquidado"] = True
                
                if existente:
                    await db.liquidaciones.update_one({"codigo_siniestro": codigo}, {"$set": liquidacion_data})
                else:
                    liquidacion_data["created_at"] = fecha_pago
                    await db.liquidaciones.insert_one(liquidacion_data)
                
                # Actualizar la orden
                await db.ordenes.update_one(
                    {"id": orden["id"]},
                    {"$set": {
                        "liquidacion_registrada": True,
                        "liquidacion_fecha": fecha_pago,
                        "liquidacion_mes": mes,
                        "updated_at": fecha_pago
                    }}
                )
                
                auto_liquidados.append({
                    "codigo": codigo, "importe": importe,
                    "numero_orden": orden.get("numero_orden"),
                    "cliente": orden.get("cliente_nombre")
                })
                    
            except Exception as e:
                errores.append({"fila": row_idx, "error": str(e)})
        
        total_liquidado = sum(i["importe"] for i in auto_liquidados)
        total_pendiente = sum(i["importe"] for i in pendientes_garantia) + sum(i["importe"] for i in no_encontrados)
        
        return {
            "success": True,
            "mes": mes,
            "resumen": {
                "auto_liquidados": len(auto_liquidados),
                "pendientes_garantia": len(pendientes_garantia),
                "ya_pagados": len(ya_pagados),
                "no_encontrados": len(no_encontrados),
                "errores": len(errores),
                "total_liquidado": round(total_liquidado, 2),
                "total_pendiente": round(total_pendiente, 2)
            },
            "auto_liquidados": auto_liquidados,
            "pendientes_garantia": pendientes_garantia,
            "ya_pagados": ya_pagados,
            "no_encontrados": no_encontrados,
            "detalles_errores": errores[:10] if errores else []
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importando liquidación: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/marcar-pagados")
async def marcar_pagados_masivo(
    codigos: List[str],
    mes: str,
    user: dict = Depends(require_master)
):
    """Marca múltiples siniestros como pagados"""
    try:
        fecha_pago = datetime.now(timezone.utc).isoformat()
        
        result = await db.liquidaciones.update_many(
            {"codigo_siniestro": {"$in": codigos}},
            {"$set": {
                "estado": "pagado",
                "fecha_pago": fecha_pago,
                "mes_liquidacion": mes,
                "updated_at": fecha_pago
            }}
        )
        
        return {
            "success": True,
            "marcados": result.modified_count,
            "fecha_pago": fecha_pago
        }
        
    except Exception as e:
        logger.error(f"Error marcando pagados: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/historial-meses")
async def obtener_historial_meses(user: dict = Depends(require_master)):
    """Obtiene resumen de liquidaciones por mes"""
    try:
        pipeline = [
            {
                "$group": {
                    "_id": "$mes_liquidacion",
                    "total_siniestros": {"$sum": 1},
                    "total_importe": {"$sum": "$importe_presupuesto"},
                    "total_garantias": {"$sum": "$costes_garantia"},
                    "pagados": {"$sum": {"$cond": [{"$eq": ["$estado", "pagado"]}, 1, 0]}},
                    "pendientes": {"$sum": {"$cond": [{"$eq": ["$estado", "pendiente"]}, 1, 0]}},
                    "reclamados": {"$sum": {"$cond": [{"$in": ["$estado", ["reclamado", "en_resolucion"]]}, 1, 0]}}
                }
            },
            {"$sort": {"_id": -1}},
            {"$limit": 12}
        ]
        
        meses = await db.liquidaciones.aggregate(pipeline).to_list(12)
        
        return {
            "meses": [
                {
                    "mes": m["_id"],
                    "total_siniestros": m["total_siniestros"],
                    "total_importe": round(m["total_importe"], 2),
                    "total_garantias": round(m["total_garantias"], 2),
                    "importe_neto": round(m["total_importe"] - m["total_garantias"], 2),
                    "pagados": m["pagados"],
                    "pendientes": m["pendientes"],
                    "reclamados": m["reclamados"]
                }
                for m in meses if m["_id"]
            ]
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/impagados")
async def obtener_impagados(user: dict = Depends(require_master)):
    """
    Obtiene siniestros con más de 60 días sin pagar.
    Genera alertas para seguimiento.
    """
    try:
        fecha_limite = (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()
        
        impagados = await db.liquidaciones.find(
            {
                "estado": "pendiente",
                "fecha_cierre": {"$lt": fecha_limite}
            },
            {"_id": 0}
        ).sort("fecha_cierre", 1).to_list(100)
        
        # Calcular días de retraso
        for imp in impagados:
            if imp.get("fecha_cierre"):
                try:
                    fecha_cierre = datetime.fromisoformat(imp["fecha_cierre"].replace("Z", "+00:00"))
                    dias = (datetime.now(timezone.utc) - fecha_cierre).days
                    imp["dias_retraso"] = dias
                except:
                    imp["dias_retraso"] = 60
        
        return {
            "impagados": impagados,
            "total": len(impagados),
            "importe_total": sum(i.get("importe_presupuesto", 0) for i in impagados)
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo impagados: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{codigo_siniestro}")
async def eliminar_liquidacion(codigo_siniestro: str, user: dict = Depends(require_master)):
    """Elimina un registro de liquidación"""
    try:
        result = await db.liquidaciones.delete_one({"codigo_siniestro": codigo_siniestro})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Liquidación no encontrada")
        
        return {"success": True, "codigo": codigo_siniestro}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error eliminando liquidación: {e}")
        raise HTTPException(status_code=500, detail=str(e))
